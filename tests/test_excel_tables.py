"""Excel 조회 테이블 변환 테스트"""

import pytest
from openpyxl import Workbook

# 테스트 대상 모듈 import
from src.utils.excel_tables import (
    convert_rula_table_a,
    get_rula_table_b,
    get_rula_table_c,
    convert_reba_table_a,
    convert_reba_table_b,
    get_reba_table_c,
    convert_owas_ac_table,
    create_rula_sheets,
    create_reba_sheets,
    create_owas_sheet,
    create_all_lookup_sheets,
)

# 원본 테이블 import (검증용)
from src.core.ergonomic.rula_calculator import RULACalculator
from src.core.ergonomic.reba_calculator import REBACalculator
from src.core.ergonomic.owas_calculator import OWASCalculator


class TestRULATableA:
    """RULA Table A 변환 테스트"""

    def test_rula_table_a_dimensions(self):
        """RULA Table A 변환 결과 차원 검증: 72행 × 2열"""
        table_2d = convert_rula_table_a()

        assert len(table_2d) == 72, f"Expected 72 rows, got {len(table_2d)}"
        assert all(len(row) == 2 for row in table_2d), "All rows should have 2 columns"

    def test_rula_table_a_conversion_first_case(self):
        """RULA Table A 변환 검증: upper_arm=1, lower_arm=1, wrist=1, wrist_twist=1"""
        table_2d = convert_rula_table_a()

        # 원본: TABLE_A[0][0][0][0] = 1
        # 변환: row = (1-1)*12 + (1-1)*4 + (1-1) = 0, col = 0
        assert table_2d[0][0] == 1

    def test_rula_table_a_conversion_middle_case(self):
        """RULA Table A 변환 검증: upper_arm=3, lower_arm=2, wrist=3, wrist_twist=2"""
        table_2d = convert_rula_table_a()

        # 원본: TABLE_A[2][1][2][1] = 4
        # 변환: row = 2*12 + 1*4 + 2 = 30, col = 1
        assert table_2d[30][1] == 4

    def test_rula_table_a_conversion_last_case(self):
        """RULA Table A 변환 검증: upper_arm=6, lower_arm=3, wrist=4, wrist_twist=2"""
        table_2d = convert_rula_table_a()

        # 원본: TABLE_A[5][2][3][1] = 9
        # 변환: row = 5*12 + 2*4 + 3 = 71, col = 1
        assert table_2d[71][1] == 9

    def test_rula_table_a_all_values_match(self):
        """RULA Table A 모든 값이 원본과 일치하는지 검증"""
        table_2d = convert_rula_table_a()
        original = RULACalculator.TABLE_A

        for ua in range(6):  # upper_arm 1-6
            for la in range(3):  # lower_arm 1-3
                for w in range(4):  # wrist 1-4
                    for wt in range(2):  # wrist_twist 1-2
                        expected = original[ua][la][w][wt]
                        row_idx = ua * 12 + la * 4 + w
                        col_idx = wt
                        actual = table_2d[row_idx][col_idx]
                        assert actual == expected, \
                            f"Mismatch at ua={ua+1}, la={la+1}, w={w+1}, wt={wt+1}: expected {expected}, got {actual}"


class TestRULATableB:
    """RULA Table B 테스트"""

    def test_rula_table_b_dimensions(self):
        """RULA Table B 차원 검증: 6행 × 12열"""
        table_2d = get_rula_table_b()

        assert len(table_2d) == 6, f"Expected 6 rows, got {len(table_2d)}"
        assert all(len(row) == 12 for row in table_2d), "All rows should have 12 columns"

    def test_rula_table_b_first_case(self):
        """RULA Table B 검증: neck=1, trunk=1, leg=1"""
        table_2d = get_rula_table_b()

        # 원본: TABLE_B[0][0][0] = 1
        # 변환: row=0, col=(1-1)*2+(1-1)=0
        assert table_2d[0][0] == 1

    def test_rula_table_b_all_values_match(self):
        """RULA Table B 모든 값이 원본과 일치하는지 검증"""
        table_2d = get_rula_table_b()
        original = RULACalculator.TABLE_B

        for n in range(6):  # neck 1-6
            for t in range(6):  # trunk 1-6
                for l in range(2):  # leg 1-2
                    expected = original[n][t][l]
                    row_idx = n
                    col_idx = t * 2 + l
                    actual = table_2d[row_idx][col_idx]
                    assert actual == expected, \
                        f"Mismatch at neck={n+1}, trunk={t+1}, leg={l+1}: expected {expected}, got {actual}"


class TestRULATableC:
    """RULA Table C 테스트"""

    def test_rula_table_c_dimensions(self):
        """RULA Table C 차원 검증: 8행 × 7열"""
        table_2d = get_rula_table_c()

        assert len(table_2d) == 8, f"Expected 8 rows, got {len(table_2d)}"
        assert all(len(row) == 7 for row in table_2d), "All rows should have 7 columns"

    def test_rula_table_c_matches_original(self):
        """RULA Table C가 원본과 일치하는지 검증"""
        table_2d = get_rula_table_c()
        original = RULACalculator.TABLE_C

        for a in range(8):
            for b in range(7):
                assert table_2d[a][b] == original[a][b], \
                    f"Mismatch at a={a+1}, b={b+1}"


class TestREBATableA:
    """REBA Table A 변환 테스트"""

    def test_reba_table_a_dimensions(self):
        """REBA Table A 변환 결과 차원 검증: 15행 × 4열"""
        table_2d = convert_reba_table_a()

        assert len(table_2d) == 15, f"Expected 15 rows, got {len(table_2d)}"
        assert all(len(row) == 4 for row in table_2d), "All rows should have 4 columns"

    def test_reba_table_a_first_case(self):
        """REBA Table A 변환 검증: neck=1, trunk=1, legs=1"""
        table_2d = convert_reba_table_a()

        # 원본: TABLE_A[0][0][0] = 1
        # 변환: row = (1-1)*5 + (1-1) = 0, col = 0
        assert table_2d[0][0] == 1

    def test_reba_table_a_all_values_match(self):
        """REBA Table A 모든 값이 원본과 일치하는지 검증"""
        table_2d = convert_reba_table_a()
        original = REBACalculator.TABLE_A

        for n in range(3):  # neck 1-3
            for t in range(5):  # trunk 1-5
                for l in range(4):  # legs 1-4
                    expected = original[n][t][l]
                    row_idx = n * 5 + t
                    col_idx = l
                    actual = table_2d[row_idx][col_idx]
                    assert actual == expected, \
                        f"Mismatch at neck={n+1}, trunk={t+1}, legs={l+1}: expected {expected}, got {actual}"


class TestREBATableB:
    """REBA Table B 변환 테스트"""

    def test_reba_table_b_dimensions(self):
        """REBA Table B 변환 결과 차원 검증: 18행 × 2열"""
        table_2d = convert_reba_table_b()

        assert len(table_2d) == 18, f"Expected 18 rows, got {len(table_2d)}"
        assert all(len(row) == 2 for row in table_2d), "All rows should have 2 columns"

    def test_reba_table_b_first_case(self):
        """REBA Table B 변환 검증: upper_arm=1, lower_arm=1, wrist=1"""
        table_2d = convert_reba_table_b()

        # 원본: TABLE_B[0][0][0] = 1
        # 변환: row = (1-1)*3 + (1-1) = 0, col = 0
        assert table_2d[0][0] == 1

    def test_reba_table_b_all_values_match(self):
        """REBA Table B 모든 값이 원본과 일치하는지 검증"""
        table_2d = convert_reba_table_b()
        original = REBACalculator.TABLE_B

        for ua in range(6):  # upper_arm 1-6
            for la in range(3):  # lower_arm 1-3
                for w in range(2):  # wrist 1-2
                    expected = original[ua][la][w]
                    row_idx = ua * 3 + la
                    col_idx = w
                    actual = table_2d[row_idx][col_idx]
                    assert actual == expected, \
                        f"Mismatch at ua={ua+1}, la={la+1}, w={w+1}: expected {expected}, got {actual}"


class TestREBATableC:
    """REBA Table C 테스트"""

    def test_reba_table_c_dimensions(self):
        """REBA Table C 차원 검증: 12행 × 12열"""
        table_2d = get_reba_table_c()

        assert len(table_2d) == 12, f"Expected 12 rows, got {len(table_2d)}"
        assert all(len(row) == 12 for row in table_2d), "All rows should have 12 columns"

    def test_reba_table_c_matches_original(self):
        """REBA Table C가 원본과 일치하는지 검증"""
        table_2d = get_reba_table_c()
        original = REBACalculator.TABLE_C

        for a in range(12):
            for b in range(12):
                assert table_2d[a][b] == original[a][b], \
                    f"Mismatch at a={a+1}, b={b+1}"


class TestOWASACTable:
    """OWAS AC Table 변환 테스트"""

    def test_owas_ac_table_dimensions(self):
        """OWAS AC Table 변환 결과 차원 검증: 12행 × 7열"""
        table_2d = convert_owas_ac_table()

        assert len(table_2d) == 12, f"Expected 12 rows, got {len(table_2d)}"
        assert all(len(row) == 7 for row in table_2d), "All rows should have 7 columns"

    def test_owas_ac_table_first_case(self):
        """OWAS AC Table 변환 검증: back=1, arms=1, legs=1"""
        table_2d = convert_owas_ac_table()

        # 원본: ACTION_CATEGORY_TABLE[(1,1,1)] = 1
        # 변환: row = (1-1)*3 + (1-1) = 0, col = 0
        assert table_2d[0][0] == 1

    def test_owas_ac_table_last_case(self):
        """OWAS AC Table 변환 검증: back=4, arms=3, legs=7"""
        table_2d = convert_owas_ac_table()

        # 원본: ACTION_CATEGORY_TABLE[(4,3,7)] = 4
        # 변환: row = 3*3 + 2 = 11, col = 6
        assert table_2d[11][6] == 4

    def test_owas_ac_table_all_values_match(self):
        """OWAS AC Table 모든 값이 원본과 일치하는지 검증"""
        table_2d = convert_owas_ac_table()
        original = OWASCalculator.ACTION_CATEGORY_TABLE

        for b in range(1, 5):  # back 1-4
            for a in range(1, 4):  # arms 1-3
                for l in range(1, 8):  # legs 1-7
                    expected = original.get((b, a, l), 1)
                    row_idx = (b - 1) * 3 + (a - 1)
                    col_idx = l - 1
                    actual = table_2d[row_idx][col_idx]
                    assert actual == expected, \
                        f"Mismatch at back={b}, arms={a}, legs={l}: expected {expected}, got {actual}"


class TestExcelSheetCreation:
    """Excel 시트 생성 테스트"""

    def test_create_rula_sheets(self):
        """RULA 조회 테이블 시트 생성 검증"""
        wb = Workbook()
        create_rula_sheets(wb)

        # 시트 존재 확인
        assert 'RULA_A' in wb.sheetnames
        assert 'RULA_B' in wb.sheetnames
        assert 'RULA_C' in wb.sheetnames

    def test_create_reba_sheets(self):
        """REBA 조회 테이블 시트 생성 검증"""
        wb = Workbook()
        create_reba_sheets(wb)

        # 시트 존재 확인
        assert 'REBA_A' in wb.sheetnames
        assert 'REBA_B' in wb.sheetnames
        assert 'REBA_C' in wb.sheetnames

    def test_create_owas_sheet(self):
        """OWAS 조회 테이블 시트 생성 검증"""
        wb = Workbook()
        create_owas_sheet(wb)

        # 시트 존재 확인
        assert 'OWAS_AC' in wb.sheetnames

    def test_create_all_lookup_sheets(self):
        """모든 조회 테이블 시트 생성 검증"""
        wb = Workbook()
        create_all_lookup_sheets(wb)

        # 모든 시트 존재 확인
        expected_sheets = ['RULA_A', 'RULA_B', 'RULA_C', 'REBA_A', 'REBA_B', 'REBA_C', 'OWAS_AC']
        for sheet_name in expected_sheets:
            assert sheet_name in wb.sheetnames, f"Sheet {sheet_name} not found"


class TestExcelNamedRanges:
    """Excel Named Range 테스트"""

    def test_rula_named_ranges(self):
        """RULA Named Range 정의 검증"""
        wb = Workbook()
        create_rula_sheets(wb)

        # Named Range 확인 (openpyxl 3.x API)
        defined_names = list(wb.defined_names)
        assert 'RULA_A' in defined_names
        assert 'RULA_B' in defined_names
        assert 'RULA_C' in defined_names

    def test_reba_named_ranges(self):
        """REBA Named Range 정의 검증"""
        wb = Workbook()
        create_reba_sheets(wb)

        # Named Range 확인
        defined_names = list(wb.defined_names)
        assert 'REBA_A' in defined_names
        assert 'REBA_B' in defined_names
        assert 'REBA_C' in defined_names

    def test_owas_named_range(self):
        """OWAS Named Range 정의 검증"""
        wb = Workbook()
        create_owas_sheet(wb)

        # Named Range 확인
        defined_names = list(wb.defined_names)
        assert 'OWAS_AC' in defined_names

    def test_all_named_ranges(self):
        """모든 Named Range 정의 검증"""
        wb = Workbook()
        create_all_lookup_sheets(wb)

        expected_names = ['RULA_A', 'RULA_B', 'RULA_C', 'REBA_A', 'REBA_B', 'REBA_C', 'OWAS_AC']
        defined_names = list(wb.defined_names)

        for name in expected_names:
            assert name in defined_names, f"Named range {name} not found"


class TestExcelSheetData:
    """Excel 시트 데이터 검증 테스트"""

    def test_rula_a_sheet_data(self):
        """RULA_A 시트 데이터 검증"""
        wb = Workbook()
        create_rula_sheets(wb)

        ws = wb['RULA_A']
        table_2d = convert_rula_table_a()

        # 데이터 검증 (일부 셀)
        assert ws.cell(row=1, column=1).value == table_2d[0][0]
        assert ws.cell(row=31, column=2).value == table_2d[30][1]
        assert ws.cell(row=72, column=2).value == table_2d[71][1]

    def test_rula_b_sheet_data(self):
        """RULA_B 시트 데이터 검증"""
        wb = Workbook()
        create_rula_sheets(wb)

        ws = wb['RULA_B']
        table_2d = get_rula_table_b()

        # 데이터 검증
        assert ws.cell(row=1, column=1).value == table_2d[0][0]
        assert ws.cell(row=6, column=12).value == table_2d[5][11]

    def test_rula_c_sheet_data(self):
        """RULA_C 시트 데이터 검증"""
        wb = Workbook()
        create_rula_sheets(wb)

        ws = wb['RULA_C']
        table_2d = get_rula_table_c()

        # 데이터 검증
        assert ws.cell(row=1, column=1).value == table_2d[0][0]
        assert ws.cell(row=8, column=7).value == table_2d[7][6]

    def test_owas_ac_sheet_data(self):
        """OWAS_AC 시트 데이터 검증"""
        wb = Workbook()
        create_owas_sheet(wb)

        ws = wb['OWAS_AC']
        table_2d = convert_owas_ac_table()

        # 데이터 검증
        assert ws.cell(row=1, column=1).value == table_2d[0][0]
        assert ws.cell(row=12, column=7).value == table_2d[11][6]
