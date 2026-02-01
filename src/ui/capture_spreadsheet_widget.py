"""
캡처 스프레드시트 위젯

캡처된 인체공학적 평가 결과를 스프레드시트 형태로 표시.
수동 입력 컬럼 편집 및 자동 재계산 기능 포함.
"""

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QAbstractItemView, QMenu, QMessageBox, QFileDialog,
    QSpinBox, QStyledItemDelegate, QLabel, QDialog,
    QComboBox, QDialogButtonBox, QFormLayout, QCheckBox,
)
from PyQt6.QtCore import Qt, pyqtSignal, QSize
from PyQt6.QtGui import QColor, QBrush, QAction, QPixmap, QImage, QIcon
from typing import List, Dict, Any, Optional
from datetime import datetime
from pathlib import Path
import json
import os


def _get_icon_path(icon_name: str) -> str:
    """아이콘 경로 반환"""
    return str(Path(__file__).parent.parent / "resources" / "icons" / f"{icon_name}.svg")

from ..core.capture_model import CaptureRecord, CaptureDataModel
from ..utils.excel_tables import create_all_lookup_sheets
from ..utils.config import Config
from ..core.logger import get_logger
from ..license import LicenseManager
from ..utils.excel_formulas import (
    get_rula_score_a_formula,
    get_rula_score_b_formula,
    get_rula_final_formula,
    get_rula_risk_formula,
    get_reba_score_a_formula,
    get_reba_score_b_formula,
    get_reba_final_formula,
    get_reba_risk_formula,
    get_owas_code_formula,
    get_owas_ac_formula,
    get_owas_risk_formula,
    get_nle_risk_formula,
    get_si_score_formula,
    get_si_risk_formula,
    # 세부 항목 합산 수식
    get_rula_upper_arm_total_formula,
    get_rula_lower_arm_total_formula,
    get_rula_wrist_total_formula,
    get_rula_neck_total_formula,
    get_rula_trunk_total_formula,
    get_reba_neck_total_formula,
    get_reba_trunk_total_formula,
    get_reba_leg_total_formula,
    get_reba_upper_arm_total_formula,
    get_reba_wrist_total_formula,
)


# =============================================================================
# 썸네일 컬럼 정의
# =============================================================================

THUMBNAIL_SIZE = 40  # 썸네일 크기 (정사각형)

# 썸네일 컬럼 (맨 앞에 추가)
THUMBNAIL_COLUMNS = [
    ('video_frame_path', 'Frame', 'thumbnail'),
    ('skeleton_image_path', 'Skeleton', 'thumbnail'),
]


# =============================================================================
# 컬럼 정의
# =============================================================================

# 컬럼 정보: (필드명, 헤더, 그룹, 편집가능, 값범위, 툴팁)
COLUMN_DEFINITIONS = [
    # 기본 정보 (3개)
    ('timestamp', 'Time', 'info', False, None, '동영상 시간\n캡처된 프레임의 타임스탬프'),
    ('frame_number', 'Frame', 'info', False, None, '프레임 번호\n동영상의 프레임 인덱스'),
    ('capture_time', 'Captured', 'info', False, None, '캡처 시각\n실제 캡처된 시간'),

    # RULA 부위 (7개)
    ('rula_upper_arm', 'Upper Arm', 'rula_body', False, None,
     '상박 총점\n= 기본 + 어깨올림 + 외전 + 팔지지\n어깨 굴곡/신전 각도에 따른 점수'),
    ('rula_lower_arm', 'Lower Arm', 'rula_body', False, None,
     '하박 총점\n= 기본 + 중앙선교차\n팔꿈치 굴곡 각도에 따른 점수'),
    ('rula_wrist', 'Wrist', 'rula_body', False, None,
     '손목 총점\n= 기본 + 측면꺾임\n손목 굴곡/신전 각도에 따른 점수'),
    ('rula_wrist_twist', 'Wrist Twist', 'rula_body', False, None,
     '손목 비틀림\n손목 회전 정도 (1-2)'),
    ('rula_neck', 'Neck', 'rula_body', False, None,
     '목 총점\n= 기본 + 회전 + 측굴\n목 굴곡/신전 각도에 따른 점수'),
    ('rula_trunk', 'Trunk', 'rula_body', False, None,
     '몸통 총점\n= 기본 + 회전 + 측굴\n몸통 굴곡 각도에 따른 점수'),
    ('rula_leg', 'Leg', 'rula_body', False, None,
     '다리 점수\n다리 지지 상태 (1-2)'),

    # RULA 수동 입력 (4개)
    ('rula_muscle_use_a', 'Muscle A', 'rula_manual', True, (0, 1),
     '근육 사용 A\n정적/반복 자세 유지 시 +1'),
    ('rula_force_load_a', 'Force A', 'rula_manual', True, (0, 3),
     '힘/하중 A\n0: 2kg 미만\n1: 2-10kg 간헐적\n2: 2-10kg 정적/반복\n3: 10kg+ 또는 충격'),
    ('rula_muscle_use_b', 'Muscle B', 'rula_manual', True, (0, 1),
     '근육 사용 B\n정적/반복 자세 유지 시 +1'),
    ('rula_force_load_b', 'Force B', 'rula_manual', True, (0, 3),
     '힘/하중 B\n0: 2kg 미만\n1: 2-10kg 간헐적\n2: 2-10kg 정적/반복\n3: 10kg+ 또는 충격'),

    # RULA 결과 (4개)
    ('rula_score_a', 'Score A', 'rula_result', False, None,
     'Wrist & Arm Score\n= Table A + Muscle A + Force A'),
    ('rula_score_b', 'Score B', 'rula_result', False, None,
     'Neck Trunk Leg Score\n= Table B + Muscle B + Force B'),
    ('rula_score', 'Score', 'rula_result', False, None,
     'RULA 최종 점수\n= Table C(Score A, Score B)'),
    ('rula_risk', 'Risk', 'rula_result', False, None,
     'RULA 위험 수준\n1-2: 허용 가능\n3-4: 추가 조사 필요\n5-6: 빠른 개선 필요\n7: 즉시 개선 필요'),

    # REBA 부위 (6개)
    ('reba_neck', 'Neck', 'reba_body', False, None,
     '목 총점\n= 기본 + 회전/측굴\n목 굴곡 각도에 따른 점수'),
    ('reba_trunk', 'Trunk', 'reba_body', False, None,
     '몸통 총점\n= 기본 + 회전/측굴\n몸통 굴곡 각도에 따른 점수'),
    ('reba_leg', 'Leg', 'reba_body', False, None,
     '다리 총점\n= 기본 + 무릎굴곡\n다리 지지 상태 및 무릎 각도'),
    ('reba_upper_arm', 'Upper Arm', 'reba_body', False, None,
     '상완 총점\n= 기본 + 어깨올림 + 외전 + 팔지지\n어깨 굴곡 각도에 따른 점수'),
    ('reba_lower_arm', 'Lower Arm', 'reba_body', False, None,
     '하완 점수\n팔꿈치 굴곡 각도 (1-2)'),
    ('reba_wrist', 'Wrist', 'reba_body', False, None,
     '손목 총점\n= 기본 + 비틀림\n손목 굴곡 각도에 따른 점수'),

    # REBA 수동 입력 (3개)
    ('reba_load_force', 'Load', 'reba_manual', True, (0, 3),
     '하중/힘\n0: 5kg 미만\n1: 5-10kg\n2: 10kg 초과\n3: 충격 또는 급격한 힘'),
    ('reba_coupling', 'Coupling', 'reba_manual', True, (0, 3),
     '커플링\n0: 좋음 (손잡이)\n1: 보통 (허용 가능)\n2: 나쁨 (어려움)\n3: 부적합'),
    ('reba_activity', 'Activity', 'reba_manual', True, (0, 3),
     '활동 점수\n정적 자세, 반복 동작,\n급격한 자세 변화 각 +1'),

    # REBA 결과 (4개)
    ('reba_score_a', 'Score A', 'reba_result', False, None,
     'Score A\n= Table A + Load/Force'),
    ('reba_score_b', 'Score B', 'reba_result', False, None,
     'Score B\n= Table B + Coupling'),
    ('reba_score', 'Score', 'reba_result', False, None,
     'REBA 최종 점수\n= Table C + Activity'),
    ('reba_risk', 'Risk', 'reba_result', False, None,
     'REBA 위험 수준\n1: 무시 가능\n2-3: 낮음\n4-7: 중간\n8-10: 높음\n11+: 매우 높음'),

    # OWAS 부위 (3개)
    ('owas_back', 'Back', 'owas_body', False, None,
     '등 자세 코드\n1: 곧은\n2: 굽힌\n3: 비튼\n4: 굽히고 비튼'),
    ('owas_arms', 'Arms', 'owas_body', False, None,
     '팔 자세 코드\n1: 양팔 어깨 아래\n2: 한팔 어깨 위\n3: 양팔 어깨 위'),
    ('owas_legs', 'Legs', 'owas_body', False, None,
     '다리 자세 코드\n1: 앉음\n2: 양다리 펴고 서기\n3: 한다리로 서기\n4: 양무릎 굽힘\n5: 한무릎 굽힘\n6: 무릎 꿇음\n7: 이동'),

    # OWAS 수동 입력 (1개)
    ('owas_load', 'Load', 'owas_manual', True, (1, 3),
     '하중 코드\n1: 10kg 미만\n2: 10-20kg\n3: 20kg 초과'),

    # OWAS 결과 (3개)
    ('owas_code', 'Code', 'owas_result', False, None,
     '자세 코드\n등+팔+다리+하중 (4자리)'),
    ('owas_ac', 'AC', 'owas_result', False, None,
     'Action Category\n1: 정상\n2: 약간 유해\n3: 명백히 유해\n4: 매우 유해'),
    ('owas_risk', 'Risk', 'owas_result', False, None,
     'OWAS 위험 수준\nAC 1: 정상\nAC 2: 약간 유해\nAC 3: 명백히 유해\nAC 4: 매우 유해'),

    # NLE 입력 (7개)
    ('nle_h', 'H (cm)', 'nle_input', True, (25, 63),
     '수평 거리 (H)\n손과 발목 사이 거리 (cm)\n범위: 25-63cm'),
    ('nle_v', 'V (cm)', 'nle_input', True, (0, 175),
     '수직 위치 (V)\n손의 바닥으로부터 높이 (cm)\n범위: 0-175cm'),
    ('nle_d', 'D (cm)', 'nle_input', True, (25, 175),
     '수직 이동 거리 (D)\n들기 시작과 끝의 높이 차이 (cm)\n범위: 25-175cm'),
    ('nle_a', 'A (°)', 'nle_input', True, (0, 135),
     '비대칭 각도 (A)\n몸통 비틀림 각도 (°)\n범위: 0-135°'),
    ('nle_f', 'F (/min)', 'nle_input', True, (0.2, 15),
     '빈도 (F)\n분당 들기 횟수\n범위: 0.2-15회/분'),
    ('nle_c', 'Coupling', 'nle_input', True, (1, 3),
     '커플링 (C)\n1: 좋음 (손잡이)\n2: 보통\n3: 나쁨'),
    ('nle_load', 'Load (kg)', 'nle_input', True, (0, 100),
     '실제 하중\n들어 올리는 물체의 무게 (kg)'),

    # NLE 결과 (3개)
    ('nle_rwl', 'RWL', 'nle_result', False, None,
     '권장 중량 한계 (RWL)\nRecommended Weight Limit (kg)'),
    ('nle_li', 'LI', 'nle_result', False, None,
     '들기 지수 (LI)\nLifting Index = Load / RWL'),
    ('nle_risk', 'Risk', 'nle_result', False, None,
     'NLE 위험 수준\nLI ≤ 1: 안전\nLI 1-3: 증가된 위험\nLI > 3: 높은 위험'),

    # SI 입력 (6개)
    ('si_ie', 'IE', 'si_input', True, (1, 5),
     '힘의 강도 (IE)\nIntensity of Exertion\n1: 가벼움 ~ 5: 최대'),
    ('si_de', 'DE', 'si_input', True, (1, 5),
     '힘 지속시간 (DE)\nDuration of Exertion\n1: <10% ~ 5: ≥80%'),
    ('si_em', 'EM', 'si_input', True, (1, 5),
     '분당 힘 횟수 (EM)\nEfforts per Minute\n1: <4 ~ 5: ≥20'),
    ('si_hwp', 'HWP', 'si_input', True, (1, 5),
     '손/손목 자세 (HWP)\nHand/Wrist Posture\n1: 중립 ~ 5: 극심한 편향'),
    ('si_sw', 'SW', 'si_input', True, (1, 5),
     '작업 속도 (SW)\nSpeed of Work\n1: 매우 느림 ~ 5: 빠름'),
    ('si_dd', 'DD', 'si_input', True, (1, 5),
     '일일 작업시간 (DD)\nDuration per Day\n1: ≤1시간 ~ 5: ≥8시간'),

    # SI 결과 (2개)
    ('si_score', 'Score', 'si_result', False, None,
     'SI 점수\nStrain Index\n= IE×DE×EM×HWP×SW×DD'),
    ('si_risk', 'Risk', 'si_result', False, None,
     'SI 위험 수준\nSI < 3: 안전\nSI 3-7: 불확실\nSI ≥ 7: 위험'),
]

# 그룹별 색상
GROUP_COLORS = {
    'info': QColor(128, 128, 128),      # 회색
    'rula_body': QColor(70, 130, 180),  # 파랑
    'rula_manual': QColor(255, 255, 150),  # 노랑
    'rula_result': QColor(100, 149, 237),  # 연파랑
    'reba_body': QColor(60, 179, 113),  # 초록
    'reba_manual': QColor(255, 255, 150),  # 노랑
    'reba_result': QColor(144, 238, 144),  # 연초록
    'owas_body': QColor(255, 165, 0),   # 주황
    'owas_manual': QColor(255, 255, 150),  # 노랑
    'owas_result': QColor(255, 200, 100),  # 연주황
    'nle_input': QColor(90, 138, 184),  # 하늘색 (편집 가능)
    'nle_result': QColor(130, 170, 210),  # 연하늘색
    'si_input': QColor(184, 168, 90),   # 황금색 (편집 가능)
    'si_result': QColor(210, 198, 130),  # 연황금색
}

# 위험 수준별 색상
RISK_COLORS = {
    # RULA
    'acceptable': QColor(144, 238, 144),  # 연초록
    'investigate': QColor(255, 255, 150),  # 노랑
    'change_soon': QColor(255, 165, 0),   # 주황
    'change_now': QColor(255, 99, 71),    # 빨강
    # REBA
    'negligible': QColor(144, 238, 144),
    'low': QColor(200, 255, 200),
    'medium': QColor(255, 255, 150),
    'high': QColor(255, 165, 0),
    'very_high': QColor(255, 99, 71),
    # OWAS
    'normal': QColor(144, 238, 144),
    'slight': QColor(255, 255, 150),
    'harmful': QColor(255, 165, 0),
    'very_harmful': QColor(255, 99, 71),
    # NLE
    'safe': QColor(144, 238, 144),        # 연초록 (LI <= 1)
    'increased': QColor(255, 165, 0),     # 주황 (LI 1-3)
    # 'high' already defined for REBA (LI > 3)
    # SI
    # 'safe' already defined for NLE
    'uncertain': QColor(255, 165, 0),     # 주황 (SI 3-7)
    'hazardous': QColor(255, 99, 71),     # 빨강 (SI >= 7)
}

# =============================================================================
# Excel 세부 컬럼 정의 (Excel 내보내기 전용)
# =============================================================================

# 세부 컬럼 매핑: total_field -> [(detail_field, header), ...]
# 각 부위의 세부 컬럼들 (토탈 컬럼 앞에 삽입됨)
RULA_DETAIL_MAP = {
    'rula_upper_arm': [
        ('rula_upper_arm_base', 'Base'),
        ('rula_upper_arm_shoulder_raised', '+Raised'),
        ('rula_upper_arm_abducted', '+Abducted'),
        ('rula_upper_arm_supported', '-Supported'),
    ],
    'rula_lower_arm': [
        ('rula_lower_arm_base', 'Base'),
        ('rula_lower_arm_working_across', '+Across'),
    ],
    'rula_wrist': [
        ('rula_wrist_base', 'Base'),
        ('rula_wrist_bent_midline', '+Bent'),
    ],
    'rula_neck': [
        ('rula_neck_base', 'Base'),
        ('rula_neck_twisted', '+Twisted'),
        ('rula_neck_side_bending', '+Side'),
    ],
    'rula_trunk': [
        ('rula_trunk_base', 'Base'),
        ('rula_trunk_twisted', '+Twisted'),
        ('rula_trunk_side_bending', '+Side'),
    ],
}

REBA_DETAIL_MAP = {
    'reba_neck': [
        ('reba_neck_base', 'Base'),
        ('reba_neck_twist_side', '+Twist/Side'),
    ],
    'reba_trunk': [
        ('reba_trunk_base', 'Base'),
        ('reba_trunk_twist_side', '+Twist/Side'),
    ],
    'reba_leg': [
        ('reba_leg_base', 'Base'),
        ('reba_leg_knee_30_60', '+Knee30-60'),
        ('reba_leg_knee_over_60', '+Knee60+'),
    ],
    'reba_upper_arm': [
        ('reba_upper_arm_base', 'Base'),
        ('reba_upper_arm_shoulder_raised', '+Raised'),
        ('reba_upper_arm_abducted', '+Abducted'),
        ('reba_upper_arm_supported', '-Supported'),
    ],
    'reba_wrist': [
        ('reba_wrist_base', 'Base'),
        ('reba_wrist_twisted', '+Twisted'),
    ],
}

# 세부 컬럼 그룹 색상
DETAIL_GROUP_COLORS = {
    'rula_detail': QColor(200, 210, 240),   # 연파랑 (RULA 세부)
    'reba_detail': QColor(200, 235, 210),   # 연초록 (REBA 세부)
}

# Risk Level 한글 매핑 (영어 키 → 한글 표시)
RISK_LABELS = {
    # RULA Risk Level
    'acceptable': '허용 가능',
    'investigate': '추가 조사 필요',
    'change_soon': '빠른 개선 필요',
    'change_now': '즉시 개선 필요',
    # REBA Risk Level
    'negligible': '무시 가능',
    'low': '낮음',
    'medium': '중간',
    'high': '높음',
    'very_high': '매우 높음',
    # OWAS Risk Level
    'normal': '정상',
    'slight': '약간 유해',
    'harmful': '명백히 유해',
    'very_harmful': '매우 유해',
    # NLE Risk Level
    'safe': '안전',
    'increased': '증가된 위험',
    # 'high' already defined for REBA
    # SI Risk Level
    # 'safe' already defined for NLE
    'uncertain': '불확실',
    'hazardous': '위험',
}

# 버튼 스타일 (PlayerWidget과 동일)
BUTTON_STYLES = {
    'json': """
        QPushButton {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 #3a9a8a, stop:1 #2a8a7a);
            color: white;
            border: none;
            padding: 6px 14px;
            border-radius: 6px;
            font-size: 12px;
            font-weight: bold;
        }
        QPushButton:hover {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 #4aaa9a, stop:1 #3a9a8a);
        }
        QPushButton:pressed {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 #2a8a7a, stop:1 #1a7a6a);
        }
    """,
    'excel': """
        QPushButton {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 #5a7ab8, stop:1 #4a6aa8);
            color: white;
            border: none;
            padding: 6px 14px;
            border-radius: 6px;
            font-size: 12px;
            font-weight: bold;
        }
        QPushButton:hover {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 #6a8ac8, stop:1 #5a7ab8);
        }
        QPushButton:pressed {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 #4a6aa8, stop:1 #3a5a98);
        }
    """,
    'delete': """
        QPushButton {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 #c55a5a, stop:1 #b54a4a);
            color: white;
            border: none;
            padding: 6px 14px;
            border-radius: 6px;
            font-size: 12px;
            font-weight: bold;
        }
        QPushButton:hover {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 #d56a6a, stop:1 #c55a5a);
        }
        QPushButton:pressed {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 #b54a4a, stop:1 #a53a3a);
        }
    """,
}


class ImageViewerDialog(QDialog):
    """이미지 원본 보기 다이얼로그"""

    def __init__(self, image_path: str, title: str = "이미지 보기", parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)

        # 이미지 라벨
        self._image_label = QLabel()
        self._image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        if os.path.exists(image_path):
            pixmap = QPixmap(image_path)
            # 최대 크기 제한 (화면의 80%)
            max_size = 800
            if pixmap.width() > max_size or pixmap.height() > max_size:
                pixmap = pixmap.scaled(
                    max_size, max_size,
                    Qt.AspectRatioMode.KeepAspectRatio,
                    Qt.TransformationMode.SmoothTransformation
                )
            self._image_label.setPixmap(pixmap)
            self.resize(pixmap.width() + 20, pixmap.height() + 60)
        else:
            self._image_label.setText("이미지를 찾을 수 없습니다.")
            self.resize(300, 100)

        layout.addWidget(self._image_label)

        # 닫기 버튼
        close_btn = QPushButton("닫기")
        close_btn.clicked.connect(self.close)
        layout.addWidget(close_btn)


class SpinBoxDelegate(QStyledItemDelegate):
    """SpinBox 에디터를 제공하는 Delegate"""

    def __init__(self, min_val: int, max_val: int, parent=None):
        super().__init__(parent)
        self._min_val = min_val
        self._max_val = max_val

    def createEditor(self, parent, option, index):
        editor = QSpinBox(parent)
        editor.setMinimum(self._min_val)
        editor.setMaximum(self._max_val)
        return editor

    def setEditorData(self, editor, index):
        value = index.model().data(index, Qt.ItemDataRole.EditRole)
        try:
            editor.setValue(int(value) if value else 0)
        except (ValueError, TypeError):
            editor.setValue(self._min_val)

    def setModelData(self, editor, model, index):
        model.setData(index, str(editor.value()), Qt.ItemDataRole.EditRole)


class CaptureSpreadsheetWidget(QWidget):
    """캡처 스프레드시트 위젯"""

    # 시그널
    record_updated = pyqtSignal(int)  # 레코드 업데이트 시 행 인덱스 전달
    export_requested = pyqtSignal(str)  # 내보내기 요청 (파일 경로)

    def __init__(self, config: Optional[Config] = None, parent=None):
        super().__init__(parent)
        self._config = config
        self._model = CaptureDataModel()
        self._updating = False  # 재계산 중 무한 루프 방지
        self._video_name: Optional[str] = None  # 현재 동영상 파일명
        self._logger = get_logger('spreadsheet')

        self._init_ui()
        self._setup_delegates()

    def _init_ui(self):
        """UI 초기화"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(5)

        # 전체 컬럼 수 (썸네일 + 데이터)
        self._thumbnail_count = len(THUMBNAIL_COLUMNS)
        total_columns = self._thumbnail_count + len(COLUMN_DEFINITIONS)

        # 테이블 위젯
        self._table = QTableWidget()
        self._table.setColumnCount(total_columns)

        # 헤더 설정 (썸네일 + 데이터)
        headers = [col[1] for col in THUMBNAIL_COLUMNS] + [col[1] for col in COLUMN_DEFINITIONS]
        self._table.setHorizontalHeaderLabels(headers)

        # 헤더 스타일
        header = self._table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        header.setStretchLastSection(True)

        # 썸네일 컬럼 고정 크기
        for i in range(self._thumbnail_count):
            self._table.setColumnWidth(i, THUMBNAIL_SIZE + 10)

        # 선택 모드
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)

        # 컨텍스트 메뉴
        self._table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._show_context_menu)

        # 셀 변경 시그널
        self._table.cellChanged.connect(self._on_cell_changed)

        # 셀 클릭 시그널 (썸네일 클릭 처리)
        self._table.cellClicked.connect(self._on_cell_clicked)

        layout.addWidget(self._table)

        # 버튼 영역
        btn_layout = QHBoxLayout()

        self._excel_btn = QPushButton(" Excel 내보내기")
        self._excel_btn.setIcon(QIcon(_get_icon_path("excel")))
        self._excel_btn.setIconSize(QSize(14, 14))
        self._excel_btn.setStyleSheet(BUTTON_STYLES['excel'])
        self._excel_btn.clicked.connect(self._export_excel)
        btn_layout.addWidget(self._excel_btn)

        self._json_btn = QPushButton(" JSON 내보내기")
        self._json_btn.setIcon(QIcon(_get_icon_path("json")))
        self._json_btn.setIconSize(QSize(14, 14))
        self._json_btn.setStyleSheet(BUTTON_STYLES['json'])
        self._json_btn.clicked.connect(self._export_json)
        btn_layout.addWidget(self._json_btn)

        btn_layout.addStretch()

        self._clear_btn = QPushButton(" 전체 삭제")
        self._clear_btn.setIcon(QIcon(_get_icon_path("trash")))
        self._clear_btn.setIconSize(QSize(14, 14))
        self._clear_btn.setStyleSheet(BUTTON_STYLES['delete'])
        self._clear_btn.clicked.connect(self._clear_all)
        btn_layout.addWidget(self._clear_btn)

        layout.addLayout(btn_layout)

        # 헤더 배경색 설정
        self._apply_header_colors()

    def _setup_delegates(self):
        """수동 입력 컬럼에 SpinBox delegate 설정"""
        for col_idx, col_def in enumerate(COLUMN_DEFINITIONS):
            field, header, group, editable, value_range = col_def[:5]
            if editable and value_range:
                min_val, max_val = value_range
                delegate = SpinBoxDelegate(min_val, max_val, self._table)
                # 썸네일 컬럼 오프셋 적용
                self._table.setItemDelegateForColumn(col_idx + self._thumbnail_count, delegate)

    def _apply_header_colors(self):
        """헤더 배경색, 스타일 및 툴팁 적용"""
        from PyQt6.QtGui import QFont

        # 헤더 폰트 설정 (진하게, 약간 큰 사이즈)
        header_font = QFont()
        header_font.setBold(True)
        header_font.setPointSize(10)

        # 썸네일 컬럼 헤더
        thumbnail_color = QColor(180, 180, 220)  # 연보라
        for col_idx, (field, header, group) in enumerate(THUMBNAIL_COLUMNS):
            item = QTableWidgetItem(header)
            item.setBackground(QBrush(thumbnail_color))
            item.setForeground(QBrush(QColor(0, 0, 0)))  # 검은색 텍스트
            item.setFont(header_font)
            item.setToolTip("클릭하여 원본 이미지 보기")
            self._table.setHorizontalHeaderItem(col_idx, item)

        # 데이터 컬럼 헤더
        for col_idx, col_def in enumerate(COLUMN_DEFINITIONS):
            field, header, group, editable, value_range = col_def[:5]
            tooltip = col_def[5] if len(col_def) > 5 else ''
            color = GROUP_COLORS.get(group, QColor(200, 200, 200))
            item = QTableWidgetItem(header)
            item.setBackground(QBrush(color))
            item.setForeground(QBrush(QColor(0, 0, 0)))  # 검은색 텍스트
            item.setFont(header_font)
            if tooltip:
                item.setToolTip(tooltip)
            self._table.setHorizontalHeaderItem(col_idx + self._thumbnail_count, item)

    def add_record(self, record: CaptureRecord) -> int:
        """
        레코드 추가

        Args:
            record: 추가할 레코드

        Returns:
            삽입된 행 인덱스
        """
        self._updating = True

        # 모델에 추가
        row_idx = self._model.add_record(record)

        # 테이블에 행 추가
        self._table.insertRow(row_idx)
        self._update_row(row_idx)

        self._updating = False
        return row_idx

    def _update_row(self, row: int):
        """행 데이터 업데이트"""
        record = self._model.get_record(row)
        if not record:
            return

        self._updating = True

        # 행 높이 설정 (썸네일 크기에 맞춤)
        self._table.setRowHeight(row, THUMBNAIL_SIZE + 4)

        # 썸네일 컬럼 업데이트
        for col_idx, (field, header, group) in enumerate(THUMBNAIL_COLUMNS):
            image_path = getattr(record, field, None)
            self._set_thumbnail_cell(row, col_idx, image_path)

        # 데이터 컬럼 업데이트
        for col_idx, col_def in enumerate(COLUMN_DEFINITIONS):
            field, header, group, editable, value_range = col_def[:5]
            value = getattr(record, field, '')

            # 타임스탬프 포맷팅
            if field == 'timestamp':
                minutes = int(value // 60)
                seconds = value % 60
                value = f"{minutes:02d}:{seconds:06.3f}"
            elif field == 'capture_time' and isinstance(value, datetime):
                value = value.strftime('%H:%M:%S')
            # Risk 컬럼 한글 변환
            elif field in ('rula_risk', 'reba_risk', 'owas_risk', 'nle_risk', 'si_risk'):
                value = RISK_LABELS.get(str(value), str(value) if value else '')
            else:
                value = str(value) if value is not None else ''

            item = QTableWidgetItem(value)

            # 읽기 전용 설정
            if not editable:
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)

            # 배경색 설정
            color = GROUP_COLORS.get(group, QColor(255, 255, 255))

            # 위험 수준 컬럼은 위험도에 따른 색상
            if field in ('rula_risk', 'reba_risk', 'owas_risk', 'nle_risk', 'si_risk'):
                risk_value = getattr(record, field, '')
                color = RISK_COLORS.get(risk_value, color)

            item.setBackground(QBrush(color))

            # 텍스트 색상 (노랑 배경은 검정 텍스트)
            if group.endswith('_manual'):
                item.setForeground(QBrush(QColor(0, 0, 0)))
            else:
                item.setForeground(QBrush(QColor(0, 0, 0)))

            self._table.setItem(row, col_idx + self._thumbnail_count, item)

        self._updating = False

    def _set_thumbnail_cell(self, row: int, col: int, image_path: Optional[str]):
        """썸네일 셀 설정"""
        self._logger.debug(f"[썸네일] _set_thumbnail_cell 호출: row={row}, col={col}, image_path={image_path}")

        label = QLabel()
        label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        label.setFixedSize(THUMBNAIL_SIZE, THUMBNAIL_SIZE)

        if image_path and os.path.exists(image_path):
            self._logger.debug(f"[썸네일] 이미지 파일 존재: {image_path}")
            pixmap = QPixmap(image_path)
            if pixmap.isNull():
                self._logger.warning(f"[썸네일] QPixmap 로드 실패: {image_path}")
            else:
                self._logger.debug(f"[썸네일] QPixmap 로드 성공: {pixmap.width()}x{pixmap.height()}")
            pixmap = pixmap.scaled(
                THUMBNAIL_SIZE, THUMBNAIL_SIZE,
                Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation
            )
            label.setPixmap(pixmap)
            label.setToolTip(f"클릭하여 원본 보기\n{image_path}")
        else:
            if image_path:
                self._logger.warning(f"[썸네일] 이미지 파일 없음: {image_path}")
            else:
                self._logger.debug(f"[썸네일] 이미지 경로 None: row={row}, col={col}")
            label.setText("-")
            label.setStyleSheet("color: #888;")

        self._table.setCellWidget(row, col, label)

    def _on_cell_clicked(self, row: int, col: int):
        """셀 클릭 시 처리 (썸네일 클릭 시 원본 보기)"""
        if col < self._thumbnail_count:
            record = self._model.get_record(row)
            if not record:
                return

            field = THUMBNAIL_COLUMNS[col][0]
            image_path = getattr(record, field, None)

            if image_path and os.path.exists(image_path):
                title = "프레임 이미지" if col == 0 else "스켈레톤 이미지"
                dialog = ImageViewerDialog(image_path, title, self)
                dialog.exec()

    def _on_cell_changed(self, row: int, col: int):
        """셀 변경 시 재계산"""
        if self._updating:
            return

        # 썸네일 컬럼은 무시
        if col < self._thumbnail_count:
            return

        # 실제 데이터 컬럼 인덱스
        data_col = col - self._thumbnail_count
        if data_col >= len(COLUMN_DEFINITIONS):
            return

        col_def = COLUMN_DEFINITIONS[data_col]
        field, header, group, editable, value_range = col_def[:5]
        if not editable:
            return

        record = self._model.get_record(row)
        if not record:
            return

        # 새 값 가져오기
        item = self._table.item(row, col)
        if not item:
            return

        try:
            new_value = int(item.text())
            # 범위 제한
            if value_range:
                new_value = max(value_range[0], min(value_range[1], new_value))
        except ValueError:
            new_value = value_range[0] if value_range else 0

        # 레코드 업데이트
        setattr(record, field, new_value)

        # 재계산
        if group == 'rula_manual':
            record.recalculate_rula()
        elif group == 'reba_manual':
            record.recalculate_reba()
        elif group == 'owas_manual':
            record.recalculate_owas()

        # 모델 업데이트
        self._model.update_record(row, record)

        # UI 갱신
        self._update_row(row)

        # 시그널
        self.record_updated.emit(row)

    def _show_context_menu(self, pos):
        """컨텍스트 메뉴 표시"""
        row = self._table.rowAt(pos.y())
        if row < 0:
            return

        menu = QMenu(self)
        delete_action = QAction("삭제", self)
        delete_action.triggered.connect(lambda: self._delete_row(row))
        menu.addAction(delete_action)

        menu.exec(self._table.viewport().mapToGlobal(pos))

    def _delete_row(self, row: int):
        """행 삭제 (이미지 삭제 옵션 포함)"""
        record = self._model.get_record(row)
        if not record:
            return

        # 이미지 경로 확인
        frame_path = getattr(record, 'video_frame_path', None)
        skeleton_path = getattr(record, 'skeleton_image_path', None)
        has_images = (frame_path and os.path.exists(frame_path)) or \
                     (skeleton_path and os.path.exists(skeleton_path))

        # Config에서 설정 가져오기
        auto_delete = True  # 기본값
        confirm_delete = True  # 기본값
        if self._config:
            auto_delete = self._config.get("images.auto_delete_on_row_delete", True)
            confirm_delete = self._config.get("images.confirm_before_delete", True)

        # 자동 삭제 + 확인 안 함 설정인 경우
        if auto_delete and not confirm_delete:
            delete_images = has_images
        # 자동 삭제 + 확인 설정인 경우
        elif auto_delete and confirm_delete:
            if has_images:
                msg_box = QMessageBox(self)
                msg_box.setWindowTitle("행 삭제")
                msg_box.setText(f"행 {row + 1}을(를) 삭제하시겠습니까?\n\n연결된 이미지 파일도 함께 삭제됩니다.")
                yes_btn = msg_box.addButton("예", QMessageBox.ButtonRole.YesRole)
                no_btn = msg_box.addButton("아니오", QMessageBox.ButtonRole.NoRole)
                msg_box.exec()
                if msg_box.clickedButton() != yes_btn:
                    return
                delete_images = True
            else:
                msg_box = QMessageBox(self)
                msg_box.setWindowTitle("행 삭제")
                msg_box.setText(f"행 {row + 1}을(를) 삭제하시겠습니까?")
                yes_btn = msg_box.addButton("예", QMessageBox.ButtonRole.YesRole)
                no_btn = msg_box.addButton("아니오", QMessageBox.ButtonRole.NoRole)
                msg_box.exec()
                if msg_box.clickedButton() != yes_btn:
                    return
                delete_images = False
        # 자동 삭제 안 함 설정인 경우 (기존 다이얼로그 사용)
        else:
            result = self._ask_delete_options(row, has_images)
            if result is None:
                return
            delete_images = result

        # 이미지 삭제 (선택한 경우)
        if delete_images:
            if frame_path and os.path.exists(frame_path):
                try:
                    os.remove(frame_path)
                except Exception:
                    pass
            if skeleton_path and os.path.exists(skeleton_path):
                try:
                    os.remove(skeleton_path)
                except Exception:
                    pass

        # 레코드 및 테이블 행 삭제
        self._model.delete_record(row)
        self._table.removeRow(row)

    def _ask_delete_options(self, row: int, has_images: bool) -> Optional[bool]:
        """
        삭제 옵션 확인 다이얼로그

        Args:
            row: 삭제할 행 인덱스
            has_images: 이미지 파일 존재 여부

        Returns:
            None: 취소됨
            True: 이미지도 삭제
            False: 데이터만 삭제
        """
        record = self._model.get_record(row)
        timestamp = getattr(record, 'timestamp', 0) if record else 0
        minutes = int(timestamp // 60)
        seconds = timestamp % 60
        time_str = f"{minutes:02d}:{seconds:05.2f}"

        if not has_images:
            # 이미지가 없으면 단순 확인
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("행 삭제")
            msg_box.setText(f"'{time_str}' 캡처 데이터를 삭제하시겠습니까?")
            yes_btn = msg_box.addButton("예", QMessageBox.ButtonRole.YesRole)
            no_btn = msg_box.addButton("아니오", QMessageBox.ButtonRole.NoRole)
            msg_box.exec()
            if msg_box.clickedButton() == yes_btn:
                return False
            return None

        # 이미지가 있는 경우 옵션 다이얼로그
        dialog = QDialog(self)
        dialog.setWindowTitle("행 삭제")
        dialog.setModal(True)

        layout = QVBoxLayout(dialog)

        # 설명
        desc_label = QLabel(
            f"'{time_str}' 캡처 데이터를 삭제하시겠습니까?\n\n"
            "이 캡처에 연결된 이미지 파일이 있습니다.\n"
            "이미지 파일도 함께 삭제하시겠습니까?"
        )
        layout.addWidget(desc_label)

        # 버튼
        button_box = QDialogButtonBox()
        delete_with_images_btn = button_box.addButton(
            "이미지도 삭제", QDialogButtonBox.ButtonRole.AcceptRole
        )
        delete_data_only_btn = button_box.addButton(
            "데이터만 삭제", QDialogButtonBox.ButtonRole.ActionRole
        )
        cancel_btn = button_box.addButton(
            "취소", QDialogButtonBox.ButtonRole.RejectRole
        )

        # 결과 저장용
        result = {"delete_images": False, "cancelled": True}

        def on_delete_with_images():
            result["delete_images"] = True
            result["cancelled"] = False
            dialog.accept()

        def on_delete_data_only():
            result["delete_images"] = False
            result["cancelled"] = False
            dialog.accept()

        def on_cancel():
            result["cancelled"] = True
            dialog.reject()

        delete_with_images_btn.clicked.connect(on_delete_with_images)
        delete_data_only_btn.clicked.connect(on_delete_data_only)
        cancel_btn.clicked.connect(on_cancel)

        layout.addWidget(button_box)

        dialog.exec()

        if result["cancelled"]:
            return None
        return result["delete_images"]

    def _clear_all(self):
        """전체 삭제"""
        if len(self._model) == 0:
            return

        # Config에서 설정 가져오기
        auto_delete = True  # 기본값
        if self._config:
            auto_delete = self._config.get("images.auto_delete_on_row_delete", True)

        # 이미지 파일 수 확인
        image_count = 0
        for record in self._model.get_all_records():
            frame_path = getattr(record, 'video_frame_path', None)
            skeleton_path = getattr(record, 'skeleton_image_path', None)
            if frame_path and os.path.exists(frame_path):
                image_count += 1
            if skeleton_path and os.path.exists(skeleton_path):
                image_count += 1

        # 확인 메시지 구성
        if auto_delete and image_count > 0:
            message = f"모든 캡처 데이터({len(self._model)}개)를 삭제하시겠습니까?\n\n" \
                      f"연결된 이미지 파일 {image_count}개도 함께 삭제됩니다."
        else:
            message = f"모든 캡처 데이터({len(self._model)}개)를 삭제하시겠습니까?"

        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("전체 삭제")
        msg_box.setText(message)
        yes_btn = msg_box.addButton("예", QMessageBox.ButtonRole.YesRole)
        no_btn = msg_box.addButton("아니오", QMessageBox.ButtonRole.NoRole)
        msg_box.exec()

        if msg_box.clickedButton() == yes_btn:
            # 이미지 삭제 (설정에 따라)
            if auto_delete:
                for record in self._model.get_all_records():
                    frame_path = getattr(record, 'video_frame_path', None)
                    skeleton_path = getattr(record, 'skeleton_image_path', None)
                    if frame_path and os.path.exists(frame_path):
                        try:
                            os.remove(frame_path)
                        except Exception:
                            pass
                    if skeleton_path and os.path.exists(skeleton_path):
                        try:
                            os.remove(skeleton_path)
                        except Exception:
                            pass

            self._model.clear()
            self._table.setRowCount(0)

    def _export_json(self):
        """JSON 내보내기"""
        # 라이센스 체크
        if not LicenseManager.instance().check_feature('json_export'):
            QMessageBox.warning(
                self, "기능 제한",
                "JSON 내보내기는 등록 버전에서 사용할 수 있습니다.\n"
                "도움말 → 라이센스 등록 메뉴에서 등록해 주세요."
            )
            return

        if len(self._model) == 0:
            QMessageBox.warning(self, "경고", "내보낼 데이터가 없습니다.")
            return

        default_filename = self._get_default_filename("json")
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "JSON 내보내기",
            default_filename,
            "JSON Files (*.json)",
        )
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(self._model.to_json())
                QMessageBox.information(self, "완료", f"JSON 파일이 저장되었습니다:\n{file_path}")
            except Exception as e:
                QMessageBox.critical(self, "오류", f"저장 중 오류 발생:\n{str(e)}")

    def _export_excel(self):
        """Excel 내보내기"""
        # 라이센스 체크
        if not LicenseManager.instance().check_feature('excel_export'):
            QMessageBox.warning(
                self, "기능 제한",
                "Excel 내보내기는 등록 버전에서 사용할 수 있습니다.\n"
                "도움말 → 라이센스 등록 메뉴에서 등록해 주세요."
            )
            return

        if len(self._model) == 0:
            QMessageBox.warning(self, "경고", "내보낼 데이터가 없습니다.")
            return

        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(
                self,
                "오류",
                "openpyxl 패키지가 설치되지 않았습니다.\npip install openpyxl 명령어로 설치해주세요.",
            )
            return

        # 내보내기 옵션 확인
        result = self._ask_export_options()
        if result is None:
            return  # 취소됨

        include_images, img_size, row_height, col_width, include_formulas, include_details = result

        default_filename = self._get_default_filename("xlsx")
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Excel 내보내기",
            default_filename,
            "Excel Files (*.xlsx)",
        )
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Capture Data"

            # 수식 포함 시 조회 테이블 시트 생성
            if include_formulas:
                create_all_lookup_sheets(wb)

            # 이미지 컬럼 오프셋 계산
            img_col_offset = len(THUMBNAIL_COLUMNS) if include_images else 0

            # 컬럼 매핑 생성 (수식용)
            col_mapping = self._build_column_mapping(img_col_offset, include_details)

            # 이미지 헤더 작성 (포함 옵션 선택 시)
            if include_images:
                thumbnail_color = "B4B4DC"  # 연보라
                for col_idx, (field, header, group) in enumerate(THUMBNAIL_COLUMNS, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.fill = PatternFill(start_color=thumbnail_color, end_color=thumbnail_color, fill_type="solid")
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    # 이미지 컬럼 너비 설정
                    ws.column_dimensions[get_column_letter(col_idx)].width = col_width

            # Excel 컬럼 목록 생성 (세부 항목 포함 시 토탈 앞에 세부 컬럼 삽입)
            excel_columns = self._build_excel_columns(include_details)

            # 데이터 헤더 작성
            for col_idx, (field, header, group, is_detail) in enumerate(excel_columns, start=1):
                cell = ws.cell(row=1, column=col_idx + img_col_offset, value=header)

                # 배경색 (세부 컬럼은 DETAIL_GROUP_COLORS, 일반 컬럼은 GROUP_COLORS)
                if is_detail:
                    color = DETAIL_GROUP_COLORS.get(group, QColor(200, 200, 200))
                else:
                    color = GROUP_COLORS.get(group, QColor(200, 200, 200))
                fill_color = f"{color.red():02X}{color.green():02X}{color.blue():02X}"
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

                # 세부 컬럼은 좁게
                if is_detail:
                    ws.column_dimensions[get_column_letter(col_idx + img_col_offset)].width = 10

            # 데이터 작성
            for row_idx, record in enumerate(self._model.get_all_records(), start=2):
                # 이미지 삽입 (포함 옵션 선택 시)
                if include_images:
                    ws.row_dimensions[row_idx].height = row_height

                    for col_idx, (field, header, group) in enumerate(THUMBNAIL_COLUMNS, start=1):
                        image_path = getattr(record, field, None)
                        if image_path and os.path.exists(image_path):
                            try:
                                img = XLImage(image_path)
                                img.width = img_size
                                img.height = img_size
                                cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
                                ws.add_image(img, cell_ref)
                            except Exception:
                                # 이미지 로드 실패 시 빈 셀
                                pass

                # 데이터 컬럼 작성 (세부 항목 포함 시 토탈 앞에 세부 컬럼)
                for col_idx, (field, header, group, is_detail) in enumerate(excel_columns, start=1):
                    value = getattr(record, field, '')
                    excel_col = col_idx + img_col_offset

                    # 수식 적용 필드인지 확인
                    formula = None
                    if include_formulas and not is_detail:
                        formula = self._get_formula_for_field(field, row_idx, col_mapping, include_details)

                    if formula:
                        # 수식 삽입
                        cell = ws.cell(row=row_idx, column=excel_col, value=formula)
                    else:
                        # 타임스탬프 포맷팅
                        if field == 'timestamp':
                            minutes = int(value // 60)
                            seconds = value % 60
                            value = f"{minutes:02d}:{seconds:06.3f}"
                        elif field == 'capture_time' and isinstance(value, datetime):
                            value = value.strftime('%H:%M:%S')
                        # Risk 컬럼 한글 변환
                        elif field in ('rula_risk', 'reba_risk', 'owas_risk', 'nle_risk', 'si_risk'):
                            value = RISK_LABELS.get(str(value), str(value) if value else '')

                        cell = ws.cell(row=row_idx, column=excel_col, value=value)

                    # 세부 컬럼 배경색
                    if is_detail:
                        color = DETAIL_GROUP_COLORS.get(group, QColor(255, 255, 255))
                        fill_color = f"{color.red():02X}{color.green():02X}{color.blue():02X}"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    # 수동 입력 컬럼 노랑 배경 (COLUMN_DEFINITIONS에서 editable 체크)
                    elif self._is_editable_field(field):
                        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                    # 위험 수준 셀 색상
                    elif field in ('rula_risk', 'reba_risk', 'owas_risk', 'nle_risk', 'si_risk') and not include_formulas:
                        risk_key = getattr(record, field, '')
                        risk_color = RISK_COLORS.get(risk_key, QColor(255, 255, 255))
                        fill_color = f"{risk_color.red():02X}{risk_color.green():02X}{risk_color.blue():02X}"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            # 데이터 컬럼 너비 자동 조정 (세부 컬럼 제외)
            for col_idx, (field, header, group, is_detail) in enumerate(excel_columns, start=1):
                if not is_detail:
                    ws.column_dimensions[get_column_letter(col_idx + img_col_offset)].width = 12

            # 수식 포함 시 조회 테이블 시트 숨김
            if include_formulas:
                for sheet_name in ['RULA_A', 'RULA_B', 'RULA_C', 'REBA_A', 'REBA_B', 'REBA_C', 'OWAS_AC']:
                    if sheet_name in wb.sheetnames:
                        wb[sheet_name].sheet_state = 'hidden'

            wb.save(file_path)
            QMessageBox.information(self, "완료", f"Excel 파일이 저장되었습니다:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "오류", f"저장 중 오류 발생:\n{str(e)}")

    def _is_editable_field(self, field: str) -> bool:
        """필드가 수동 입력 가능한지 확인"""
        for col_def in COLUMN_DEFINITIONS:
            if col_def[0] == field:
                return col_def[3]  # editable 필드
        return False

    def _build_excel_columns(self, include_details: bool) -> list:
        """
        Excel 컬럼 목록 생성 (세부 항목 포함 시 토탈 앞에 세부 컬럼 삽입)

        Args:
            include_details: 세부 컬럼 포함 여부

        Returns:
            컬럼 정보 리스트: [(field, header, group, is_detail), ...]
        """
        columns = []
        detail_maps = {**RULA_DETAIL_MAP, **REBA_DETAIL_MAP}

        for col_def in COLUMN_DEFINITIONS:
            field = col_def[0]
            header = col_def[1]
            group = col_def[2]

            # 세부 항목이 있는 부위인지 확인
            if include_details and field in detail_maps:
                # 세부 컬럼들 먼저 추가
                detail_group = 'rula_detail' if field.startswith('rula_') else 'reba_detail'
                for detail_field, detail_header in detail_maps[field]:
                    columns.append((detail_field, detail_header, detail_group, True))
                # 토탈 컬럼 추가 (헤더에 Total 표시)
                columns.append((field, f"{header}", group, False))
            else:
                columns.append((field, header, group, False))

        return columns

    def _build_column_mapping(self, img_col_offset: int, include_details: bool = False) -> Dict[str, str]:
        """
        필드명 -> Excel 컬럼 문자 매핑 생성

        Args:
            img_col_offset: 이미지 컬럼 오프셋
            include_details: 세부 컬럼 포함 여부

        Returns:
            필드명과 Excel 컬럼 문자의 매핑 딕셔너리
        """
        from openpyxl.utils import get_column_letter

        mapping = {}
        excel_columns = self._build_excel_columns(include_details)

        for col_idx, (field, header, group, is_detail) in enumerate(excel_columns, start=1):
            excel_col = get_column_letter(col_idx + img_col_offset)
            mapping[field] = excel_col

        return mapping

    def _get_formula_for_field(
        self, field: str, row: int, col_mapping: Dict[str, str], include_details: bool = False
    ) -> Optional[str]:
        """
        필드에 해당하는 Excel 수식 반환

        Args:
            field: 필드명
            row: Excel 행 번호
            col_mapping: 필드 -> 컬럼 매핑
            include_details: 세부 컬럼 포함 여부 (True면 total 컬럼에 합산 수식 적용)

        Returns:
            Excel 수식 문자열 또는 None (수식이 아닌 필드)
        """
        # 세부 컬럼 포함 시 부위 total 컬럼에 합산 수식 적용
        if include_details:
            # RULA 부위 총점 수식
            if field == 'rula_upper_arm':
                return get_rula_upper_arm_total_formula(row, {
                    'base': col_mapping['rula_upper_arm_base'],
                    'shoulder_raised': col_mapping['rula_upper_arm_shoulder_raised'],
                    'abducted': col_mapping['rula_upper_arm_abducted'],
                    'supported': col_mapping['rula_upper_arm_supported'],
                })
            elif field == 'rula_lower_arm':
                return get_rula_lower_arm_total_formula(row, {
                    'base': col_mapping['rula_lower_arm_base'],
                    'working_across': col_mapping['rula_lower_arm_working_across'],
                })
            elif field == 'rula_wrist':
                return get_rula_wrist_total_formula(row, {
                    'base': col_mapping['rula_wrist_base'],
                    'bent_midline': col_mapping['rula_wrist_bent_midline'],
                })
            elif field == 'rula_neck':
                return get_rula_neck_total_formula(row, {
                    'base': col_mapping['rula_neck_base'],
                    'twisted': col_mapping['rula_neck_twisted'],
                    'side_bending': col_mapping['rula_neck_side_bending'],
                })
            elif field == 'rula_trunk':
                return get_rula_trunk_total_formula(row, {
                    'base': col_mapping['rula_trunk_base'],
                    'twisted': col_mapping['rula_trunk_twisted'],
                    'side_bending': col_mapping['rula_trunk_side_bending'],
                })
            # REBA 부위 총점 수식
            elif field == 'reba_neck':
                return get_reba_neck_total_formula(row, {
                    'base': col_mapping['reba_neck_base'],
                    'twist_side': col_mapping['reba_neck_twist_side'],
                })
            elif field == 'reba_trunk':
                return get_reba_trunk_total_formula(row, {
                    'base': col_mapping['reba_trunk_base'],
                    'twist_side': col_mapping['reba_trunk_twist_side'],
                })
            elif field == 'reba_leg':
                return get_reba_leg_total_formula(row, {
                    'base': col_mapping['reba_leg_base'],
                    'knee_30_60': col_mapping['reba_leg_knee_30_60'],
                    'knee_over_60': col_mapping['reba_leg_knee_over_60'],
                })
            elif field == 'reba_upper_arm':
                return get_reba_upper_arm_total_formula(row, {
                    'base': col_mapping['reba_upper_arm_base'],
                    'shoulder_raised': col_mapping['reba_upper_arm_shoulder_raised'],
                    'abducted': col_mapping['reba_upper_arm_abducted'],
                    'supported': col_mapping['reba_upper_arm_supported'],
                })
            elif field == 'reba_wrist':
                return get_reba_wrist_total_formula(row, {
                    'base': col_mapping['reba_wrist_base'],
                    'twisted': col_mapping['reba_wrist_twisted'],
                })

        # RULA 수식
        if field == 'rula_score_a':
            return get_rula_score_a_formula(row, {
                'upper_arm': col_mapping['rula_upper_arm'],
                'lower_arm': col_mapping['rula_lower_arm'],
                'wrist': col_mapping['rula_wrist'],
                'wrist_twist': col_mapping['rula_wrist_twist'],
                'muscle_a': col_mapping['rula_muscle_use_a'],
                'force_a': col_mapping['rula_force_load_a'],
            })
        elif field == 'rula_score_b':
            return get_rula_score_b_formula(row, {
                'neck': col_mapping['rula_neck'],
                'trunk': col_mapping['rula_trunk'],
                'leg': col_mapping['rula_leg'],
                'muscle_b': col_mapping['rula_muscle_use_b'],
                'force_b': col_mapping['rula_force_load_b'],
            })
        elif field == 'rula_score':
            return get_rula_final_formula(row, {
                'score_a': col_mapping['rula_score_a'],
                'score_b': col_mapping['rula_score_b'],
            })
        elif field == 'rula_risk':
            return get_rula_risk_formula(row, col_mapping['rula_score'])

        # REBA 수식
        elif field == 'reba_score_a':
            return get_reba_score_a_formula(row, {
                'neck': col_mapping['reba_neck'],
                'trunk': col_mapping['reba_trunk'],
                'leg': col_mapping['reba_leg'],
                'load': col_mapping['reba_load_force'],
            })
        elif field == 'reba_score_b':
            return get_reba_score_b_formula(row, {
                'upper_arm': col_mapping['reba_upper_arm'],
                'lower_arm': col_mapping['reba_lower_arm'],
                'wrist': col_mapping['reba_wrist'],
                'coupling': col_mapping['reba_coupling'],
            })
        elif field == 'reba_score':
            return get_reba_final_formula(row, {
                'score_a': col_mapping['reba_score_a'],
                'score_b': col_mapping['reba_score_b'],
                'activity': col_mapping['reba_activity'],
            })
        elif field == 'reba_risk':
            return get_reba_risk_formula(row, col_mapping['reba_score'])

        # OWAS 수식
        elif field == 'owas_code':
            return get_owas_code_formula(row, {
                'back': col_mapping['owas_back'],
                'arms': col_mapping['owas_arms'],
                'legs': col_mapping['owas_legs'],
                'load': col_mapping['owas_load'],
            })
        elif field == 'owas_ac':
            return get_owas_ac_formula(row, {
                'back': col_mapping['owas_back'],
                'arms': col_mapping['owas_arms'],
                'legs': col_mapping['owas_legs'],
            })
        elif field == 'owas_risk':
            return get_owas_risk_formula(row, col_mapping['owas_ac'])

        return None

    def _ask_export_options(self) -> Optional[tuple]:
        """
        Excel 내보내기 옵션 확인 다이얼로그

        Returns:
            None: 취소됨
            (include_images, img_size, row_height, col_width, include_formulas, include_details): 옵션
            - include_images: 이미지 포함 여부
            - img_size: 이미지 크기 (픽셀)
            - row_height: 행 높이 (포인트)
            - col_width: 열 너비
            - include_formulas: 수식 포함 여부
            - include_details: 세부 항목 포함 여부
        """
        # 이미지 크기 옵션: (라벨, 이미지크기, 행높이, 열너비)
        SIZE_OPTIONS = [
            ("작게 (50x50)", 50, 40, 8),
            ("보통 (150x150)", 150, 115, 22),
            ("크게 (200x200)", 200, 155, 29),
            ("매우 크게 (300x300)", 300, 230, 43),
        ]
        DEFAULT_INDEX = 1  # 보통

        dialog = QDialog(self)
        dialog.setWindowTitle("Excel 내보내기 옵션")
        dialog.setModal(True)

        layout = QVBoxLayout(dialog)

        # 이미지 옵션 섹션
        img_section_label = QLabel("<b>이미지 옵션</b>")
        layout.addWidget(img_section_label)

        img_checkbox = QCheckBox("이미지 포함")
        img_checkbox.setChecked(True)  # 기본 선택
        img_checkbox.setToolTip("Excel 파일에 캡처 이미지를 포함합니다.\n파일 크기가 증가합니다.")
        layout.addWidget(img_checkbox)

        # 이미지 크기 콤보박스 (체크박스 활성화 시만 사용)
        form_layout = QFormLayout()
        size_combo = QComboBox()
        for label, _, _, _ in SIZE_OPTIONS:
            size_combo.addItem(label)
        size_combo.setCurrentIndex(DEFAULT_INDEX)
        size_combo.setEnabled(True)  # 이미지 체크박스가 기본 선택이므로 활성화
        form_layout.addRow("이미지 크기:", size_combo)
        layout.addLayout(form_layout)

        # 이미지 체크박스와 콤보박스 연동
        img_checkbox.toggled.connect(size_combo.setEnabled)

        # 구분선
        layout.addSpacing(10)

        # 수식 옵션 섹션
        formula_section_label = QLabel("<b>수식 옵션</b>")
        layout.addWidget(formula_section_label)

        formula_checkbox = QCheckBox("자동 계산 수식 포함")
        formula_checkbox.setChecked(True)  # 기본 선택
        formula_checkbox.setToolTip(
            "결과 컬럼(Score A/B, Score, Risk)에 Excel 수식을 삽입합니다.\n"
            "입력값 변경 시 결과가 자동으로 재계산됩니다.\n"
            "조회 테이블이 별도 시트로 생성됩니다."
        )
        layout.addWidget(formula_checkbox)

        # 구분선
        layout.addSpacing(10)

        # 세부 항목 옵션 섹션
        detail_section_label = QLabel("<b>세부 항목 옵션</b>")
        layout.addWidget(detail_section_label)

        detail_checkbox = QCheckBox("RULA/REBA 세부 점수 포함")
        detail_checkbox.setChecked(True)  # 기본 선택
        detail_checkbox.setToolTip(
            "각 부위 점수의 세부 항목(Base, +Raised 등)을 포함합니다.\n"
            "세부 항목은 파일 끝에 별도 컬럼으로 추가됩니다.\n"
            "수식 포함 시 합계(Total)가 자동 계산됩니다."
        )
        layout.addWidget(detail_checkbox)

        layout.addSpacing(10)

        # 버튼
        button_box = QDialogButtonBox()
        ok_btn = button_box.addButton("확인", QDialogButtonBox.ButtonRole.AcceptRole)
        cancel_btn = button_box.addButton("취소", QDialogButtonBox.ButtonRole.RejectRole)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        if dialog.exec() != QDialog.DialogCode.Accepted:
            return None

        include_images = img_checkbox.isChecked()
        include_formulas = formula_checkbox.isChecked()
        include_details = detail_checkbox.isChecked()

        if include_images:
            idx = size_combo.currentIndex()
            _, img_size, row_height, col_width = SIZE_OPTIONS[idx]
            return (True, img_size, row_height, col_width, include_formulas, include_details)
        else:
            return (False, 0, 0, 0, include_formulas, include_details)

    def get_model(self) -> CaptureDataModel:
        """데이터 모델 반환"""
        return self._model

    def clear_all(self):
        """모든 레코드 삭제 및 테이블 초기화"""
        self._model.clear()
        self._table.setRowCount(0)

    def load_from_model(self, model: CaptureDataModel):
        """CaptureDataModel에서 데이터 로드"""
        self._logger.info(f"[썸네일] load_from_model 호출: 레코드 수={len(model)}")
        self.clear_all()
        for idx, record in enumerate(model.get_all_records()):
            self._logger.debug(f"[썸네일] 레코드 {idx}: video_frame_path={record.video_frame_path}, skeleton_image_path={record.skeleton_image_path}")
            self.add_record(record)

    def get_record_count(self) -> int:
        """레코드 수 반환"""
        return len(self._model)

    def set_video_name(self, video_name: str):
        """동영상 파일명 설정"""
        self._video_name = video_name

    def _get_default_filename(self, extension: str) -> str:
        """기본 파일명 생성 (동영상제목_타임스탬프.확장자)"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if self._video_name:
            return f"{self._video_name}_{timestamp}.{extension}"
        return f"capture_{timestamp}.{extension}"
