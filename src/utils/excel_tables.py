"""
Excel 조회 테이블 변환 모듈

다차원 Python 테이블을 Excel INDEX 함수에서 사용할 수 있는
2차원 배열로 변환합니다.
"""

from typing import List

from ..core.ergonomic.rula_calculator import RULACalculator
from ..core.ergonomic.reba_calculator import REBACalculator
from ..core.ergonomic.owas_calculator import OWASCalculator


def convert_rula_table_a() -> List[List[int]]:
    """
    RULA Table A를 4차원에서 2차원으로 변환

    원본 구조: [upper_arm 0-5][lower_arm 0-2][wrist 0-3][wrist_twist 0-1]
    변환 구조: 72행 × 2열
        - 행 인덱스 = upper_arm*12 + lower_arm*4 + wrist (0-71)
        - 열 인덱스 = wrist_twist (0-1)

    Returns:
        72행 × 2열 2차원 리스트
    """
    original = RULACalculator.TABLE_A
    table_2d = []

    for ua in range(6):  # upper_arm 0-5
        for la in range(3):  # lower_arm 0-2
            for w in range(4):  # wrist 0-3
                row = []
                for wt in range(2):  # wrist_twist 0-1
                    row.append(original[ua][la][w][wt])
                table_2d.append(row)

    return table_2d


def get_rula_table_b() -> List[List[int]]:
    """
    RULA Table B를 3차원에서 2차원으로 변환

    원본 구조: [neck 0-5][trunk 0-5][leg 0-1]
    변환 구조: 6행 × 12열
        - 행 인덱스 = neck (0-5)
        - 열 인덱스 = trunk*2 + leg (0-11)

    Returns:
        6행 × 12열 2차원 리스트
    """
    original = RULACalculator.TABLE_B
    table_2d = []

    for n in range(6):  # neck 0-5
        row = []
        for t in range(6):  # trunk 0-5
            for l in range(2):  # leg 0-1
                row.append(original[n][t][l])
        table_2d.append(row)

    return table_2d


def get_rula_table_c() -> List[List[int]]:
    """
    RULA Table C 반환 (이미 2차원)

    구조: 8행 × 7열
        - 행 인덱스 = Score A (1-8) -> 0-7
        - 열 인덱스 = Score B (1-7) -> 0-6

    Returns:
        8행 × 7열 2차원 리스트 (원본의 복사본)
    """
    return [row[:] for row in RULACalculator.TABLE_C]


def convert_reba_table_a() -> List[List[int]]:
    """
    REBA Table A를 3차원에서 2차원으로 변환

    원본 구조: [neck 0-2][trunk 0-4][legs 0-3]
    변환 구조: 15행 × 4열
        - 행 인덱스 = neck*5 + trunk (0-14)
        - 열 인덱스 = legs (0-3)

    Returns:
        15행 × 4열 2차원 리스트
    """
    original = REBACalculator.TABLE_A
    table_2d = []

    for n in range(3):  # neck 0-2
        for t in range(5):  # trunk 0-4
            row = []
            for l in range(4):  # legs 0-3
                row.append(original[n][t][l])
            table_2d.append(row)

    return table_2d


def convert_reba_table_b() -> List[List[int]]:
    """
    REBA Table B를 3차원에서 2차원으로 변환

    원본 구조: [upper_arm 0-5][lower_arm 0-2][wrist 0-1]
    변환 구조: 18행 × 2열
        - 행 인덱스 = upper_arm*3 + lower_arm (0-17)
        - 열 인덱스 = wrist (0-1)

    Returns:
        18행 × 2열 2차원 리스트
    """
    original = REBACalculator.TABLE_B
    table_2d = []

    for ua in range(6):  # upper_arm 0-5
        for la in range(3):  # lower_arm 0-2
            row = []
            for w in range(2):  # wrist 0-1
                row.append(original[ua][la][w])
            table_2d.append(row)

    return table_2d


def get_reba_table_c() -> List[List[int]]:
    """
    REBA Table C 반환 (이미 2차원)

    구조: 12행 × 12열
        - 행 인덱스 = Score A (1-12) -> 0-11
        - 열 인덱스 = Score B (1-12) -> 0-11

    Returns:
        12행 × 12열 2차원 리스트 (원본의 복사본)
    """
    return [row[:] for row in REBACalculator.TABLE_C]


def convert_owas_ac_table() -> List[List[int]]:
    """
    OWAS Action Category 딕셔너리를 2차원 배열로 변환

    원본 구조: {(back, arms, legs): ac} 딕셔너리
    변환 구조: 12행 × 7열
        - 행 인덱스 = (back-1)*3 + (arms-1) (0-11)
        - 열 인덱스 = legs-1 (0-6)

    Returns:
        12행 × 7열 2차원 리스트
    """
    original = OWASCalculator.ACTION_CATEGORY_TABLE
    table_2d = []

    for b in range(1, 5):  # back 1-4
        for a in range(1, 4):  # arms 1-3
            row = []
            for l in range(1, 8):  # legs 1-7
                row.append(original.get((b, a, l), 1))
            table_2d.append(row)

    return table_2d


def _write_table_to_sheet(ws, table_2d: List[List[int]]):
    """2차원 테이블을 시트에 작성"""
    for row_idx, row_data in enumerate(table_2d, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def _create_named_range(wb, name: str, sheet_name: str, rows: int, cols: int):
    """Named Range 생성"""
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.utils import get_column_letter

    end_col = get_column_letter(cols)
    ref = f"'{sheet_name}'!$A$1:${end_col}${rows}"
    defn = DefinedName(name=name, attr_text=ref)
    wb.defined_names.add(defn)


def create_rula_sheets(wb):
    """
    RULA 조회 테이블 시트 생성

    Args:
        wb: openpyxl Workbook 객체

    생성되는 시트:
        - RULA_A: 72행 × 2열
        - RULA_B: 6행 × 12열
        - RULA_C: 8행 × 7열
    """
    # RULA_A 시트
    ws_a = wb.create_sheet("RULA_A")
    table_a = convert_rula_table_a()
    _write_table_to_sheet(ws_a, table_a)
    _create_named_range(wb, "RULA_A", "RULA_A", 72, 2)

    # RULA_B 시트
    ws_b = wb.create_sheet("RULA_B")
    table_b = get_rula_table_b()
    _write_table_to_sheet(ws_b, table_b)
    _create_named_range(wb, "RULA_B", "RULA_B", 6, 12)

    # RULA_C 시트
    ws_c = wb.create_sheet("RULA_C")
    table_c = get_rula_table_c()
    _write_table_to_sheet(ws_c, table_c)
    _create_named_range(wb, "RULA_C", "RULA_C", 8, 7)


def create_reba_sheets(wb):
    """
    REBA 조회 테이블 시트 생성

    Args:
        wb: openpyxl Workbook 객체

    생성되는 시트:
        - REBA_A: 15행 × 4열
        - REBA_B: 18행 × 2열
        - REBA_C: 12행 × 12열
    """
    # REBA_A 시트
    ws_a = wb.create_sheet("REBA_A")
    table_a = convert_reba_table_a()
    _write_table_to_sheet(ws_a, table_a)
    _create_named_range(wb, "REBA_A", "REBA_A", 15, 4)

    # REBA_B 시트
    ws_b = wb.create_sheet("REBA_B")
    table_b = convert_reba_table_b()
    _write_table_to_sheet(ws_b, table_b)
    _create_named_range(wb, "REBA_B", "REBA_B", 18, 2)

    # REBA_C 시트
    ws_c = wb.create_sheet("REBA_C")
    table_c = get_reba_table_c()
    _write_table_to_sheet(ws_c, table_c)
    _create_named_range(wb, "REBA_C", "REBA_C", 12, 12)


def create_owas_sheet(wb):
    """
    OWAS 조회 테이블 시트 생성

    Args:
        wb: openpyxl Workbook 객체

    생성되는 시트:
        - OWAS_AC: 12행 × 7열
    """
    ws = wb.create_sheet("OWAS_AC")
    table = convert_owas_ac_table()
    _write_table_to_sheet(ws, table)
    _create_named_range(wb, "OWAS_AC", "OWAS_AC", 12, 7)


def create_si_sheets(wb):
    """
    SI (Strain Index) 조회 테이블 시트 생성

    Args:
        wb: openpyxl Workbook 객체

    생성되는 시트:
        - SI_IE: Intensity of Exertion (1x5)
        - SI_DE: Duration of Exertion (1x5)
        - SI_EM: Efforts per Minute (1x5)
        - SI_HWP: Hand/Wrist Posture (1x5)
        - SI_SW: Speed of Work (1x5)
        - SI_DD: Duration per Day (1x5)
    """
    # SI Multiplier 테이블 (레벨 1-5)
    # 소스: core/score_calculator.py SI_MULTIPLIERS
    si_tables = {
        'SI_IE': [1.0, 3.0, 6.0, 9.0, 13.0],      # Intensity of Exertion
        'SI_DE': [0.5, 1.0, 1.5, 2.0, 3.0],       # Duration of Exertion
        'SI_EM': [0.5, 1.0, 1.5, 2.0, 3.0],       # Efforts per Minute
        'SI_HWP': [1.0, 1.0, 1.5, 2.0, 3.0],      # Hand/Wrist Posture
        'SI_SW': [1.0, 1.0, 1.0, 1.5, 2.0],       # Speed of Work
        'SI_DD': [0.25, 0.5, 0.75, 1.0, 1.5],     # Duration per Day
    }

    for name, values in si_tables.items():
        ws = wb.create_sheet(name)
        # 1행에 가로로 작성 (INDEX 함수에서 1차원 배열로 사용)
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=1, column=col_idx, value=value)
        # Named Range 생성 (1행 × 5열)
        _create_named_range(wb, name, name, 1, 5)


def create_all_lookup_sheets(wb):
    """
    모든 조회 테이블 시트 생성

    Args:
        wb: openpyxl Workbook 객체

    생성되는 시트:
        - RULA_A, RULA_B, RULA_C
        - REBA_A, REBA_B, REBA_C
        - OWAS_AC
        - SI_IE, SI_DE, SI_EM, SI_HWP, SI_SW, SI_DD
    """
    create_rula_sheets(wb)
    create_reba_sheets(wb)
    create_owas_sheet(wb)
    create_si_sheets(wb)
